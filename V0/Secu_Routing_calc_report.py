import pandas as pd
import numpy as np
from tkinter import Tk, Button, Label, filedialog, messagebox
from openpyxl import load_workbook
from datetime import datetime, timedelta
from openpyxl.styles import Font, Border, Side, NamedStyle, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def select_file():
    """Open a file dialog to select a CSV file and return its path."""
    file_path = filedialog.askopenfilename(
        title="Select the CSV file",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    return file_path

def save_file():
    """Open a file dialog to specify the Excel file to save the report."""
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    return file_path

def process_data(csv_file_path, excel_path):
    """Process the CSV file and save the report to an Excel file."""
    df = pd.read_csv(csv_file_path, low_memory=False)
    df_copy = df.copy()

    # Convert dates to datetime format for comparison
    df_copy['FirstSignalDate'] = pd.to_datetime(df_copy['FirstSignalDate'], errors='coerce')
    df_copy['LastSignalDate'] = pd.to_datetime(df_copy['LastSignalDate'], errors='coerce')

    # Get the current year
    current_year = pd.Timestamp.now().year

    # Create 'DeviceActive' column
    start_date = pd.Timestamp(f'{current_year}-04-01')
    end_date = pd.Timestamp(f'{current_year}-09-30')

    df_copy['DeviceActive'] = 'Inactive'
    df_copy.loc[
        (df_copy['LastSignalDate'] > start_date) &
        (df_copy['LastSignalDate'] < end_date) &
        ((df_copy['LastSignalDate'] - df_copy['FirstSignalDate']).dt.days > 10) &
        (df_copy['ItemCode'].isin(['17300', '15300'])),
        'DeviceActive'
    ] = 'Active'

    # Create 'DaysActive' column
    df_copy['DaysActive'] = 0
    df_copy.loc[
        df_copy['DeviceActive'] == 'Active',
        'DaysActive'
    ] = df_copy.apply(
        lambda row: (row['LastSignalDate'] - row['FirstSignalDate']).days 
        if row['FirstSignalDate'] > pd.Timestamp(f'{current_year}-03-30') 
        else (row['LastSignalDate'] - start_date).days, axis=1
    )

    # Create 'MonthsActive' column
    df_copy['MonthsActive'] = (df_copy['DaysActive'] / 30).apply(lambda x: -(-x // 1))  # Ceiling division

    # Create 'Fee ex VAT' column
    df_copy['Fee ex VAT'] = df_copy['MonthsActive'] * df_copy['Amount']

    # Filter for ItemCode 17300 and 15300
    df_filtered = df_copy[df_copy['ItemCode'].isin(['17300', '15300'])].copy()

    # Create summary DataFrame
    summary_df = df_filtered.groupby('SabreCode').agg(
        Branch=('Branch', 'first'),
        ItemCode_17300=('Fee ex VAT', lambda x: x[df_filtered['ItemCode'] == '17300'].sum()),
        ItemCode_15300=('Fee ex VAT', lambda x: x[df_filtered['ItemCode'] == '15300'].sum()),
        TotalActive=('DeviceActive', lambda x: (x == 'Active').sum()),
        Total_ex_VAT=('Fee ex VAT', 'sum')
    ).reset_index()

    # Calculate 'Price Per Unit' and add it to the DataFrame
    summary_df['Price Per Unit'] = summary_df.apply(
        lambda row: row['Total_ex_VAT'] / row['TotalActive'] if row['TotalActive'] > 0 else 0,
        axis=1
    )

    # Rename columns
    summary_df.rename(columns={
        'ItemCode_17300': '17300',
        'ItemCode_15300': '15300'
    }, inplace=True)

    # Calculate the total sum for 'Total_ex_VAT'
    total_ex_vat = summary_df['Total_ex_VAT'].sum()

    # Append the total and empty row
    total_row = pd.DataFrame({
        'SabreCode': ['Total'],
        'Branch': [None],
        '17300': [None],
        '15300': [None],
        'TotalActive': [None],
        'Price Per Unit': [None],
        'Total_ex_VAT': [total_ex_vat]
    })

    empty_row = pd.DataFrame({
        'SabreCode': [None],
        'Branch': [None],
        '17300': [None],
        '15300': [None],
        'TotalActive': [None],
        'Price Per Unit': [None],
        'Total_ex_VAT': [None]
    })

    empty_row = empty_row.dropna(how='all', axis=1)
    total_row = total_row.dropna(how='all', axis=1)

    summary_df = pd.concat([summary_df, empty_row, total_row], ignore_index=True)

    # Save the summary DataFrame to Excel
    summary_df.to_excel(excel_path, index=False, sheet_name='Summary')

    # Open the workbook and select the sheet
    wb = load_workbook(excel_path)
    ws = wb['Summary']

    # Define styles
    currency_format = NamedStyle(name='currency', number_format='R #,##0.00')
    number_format = NamedStyle(name='number', number_format='0')
    bold_font = Font(bold=True)
    blue_font = Font(color="0000FF")
    red_font = Font(color="FF0000")
    header_fill = PatternFill(start_color="000099", end_color="000099", fill_type="solid")
    header_font = Font(color="FFFFFF")
    border_color = "D9D9D9"

    # Define border styles
    border_style = Border(
        left=Side(border_style='thin', color=border_color),
        right=Side(border_style='thin', color=border_color),
        top=Side(border_style='thin', color=border_color),
        bottom=Side(border_style='thin', color=border_color)
    )

    # Define 'Top and Thick Bottom' border style for the 'Total_ex_VAT' cell with black color
    total_ex_vat_border = Border(
        top=Side(border_style='thin', color="000000"),  # Black color
        bottom=Side(border_style='thick', color="000000"),  # Black color
        left=Side(border_style='none'),
        right=Side(border_style='none')
    )

    # Apply formatting to specific columns
    currency_cols = ['17300', '15300', 'Total_ex_VAT', 'Amount']
    for col in currency_cols:
        if col in summary_df.columns:
            col_letter = chr(ord('A') + summary_df.columns.get_loc(col))
            for cell in ws[col_letter][1:]:  # Skip header row
                cell.style = currency_format

    # Apply numeric formatting to 'DaysActive' and 'MonthsActive'
    numeric_cols = ['DaysActive', 'MonthsActive']
    for col in numeric_cols:
        if col in summary_df.columns:
            col_letter = chr(ord('A') + summary_df.columns.get_loc(col))
            for cell in ws[col_letter][1:]:  # Skip header row
                cell.style = number_format

    # Apply fill color and font color for headings on 'Summary' sheet
    for cell in ws[1]:  # Apply styles to the first row (headers)
        cell.fill = header_fill
        cell.font = header_font

    # Apply red font formatting to rows where 'Total_ex_VAT' = 0
    total_ex_vat_col_index = summary_df.columns.get_loc('Total_ex_VAT') + 1  # +1 for 1-based index
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        total_ex_vat_cell = row[total_ex_vat_col_index - 1]  # Indexing correction for 0-based indexing
        if total_ex_vat_cell.value is not None and total_ex_vat_cell.value == 0:
            for cell in row:
                cell.font = red_font

    # Apply borders to the headings (first row)
    for cell in ws[1]:  # First row
        cell.border = border_style

    # Apply bold font to 'Total' row
    total_row_index = summary_df.index[summary_df['SabreCode'] == 'Total'].tolist()[0] + 2  # +2 for zero-based index and header row
    total_row_cells = ws[total_row_index]

    # Apply bold font to 'Total' row
    for cell in total_row_cells:
        cell.font = bold_font

    # Apply bold font to 'Total_ex_VAT' values greater than 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=total_ex_vat_col_index, max_col=total_ex_vat_col_index):
        total_ex_vat_cell = row[0]
        if total_ex_vat_cell.value is not None and total_ex_vat_cell.value > 0:
            total_ex_vat_cell.font = bold_font

    # Apply 'Top and Thick Bottom' border style with black color to the 'Total_ex_VAT' cell in the 'Total' row
    total_ex_vat_total_cell = ws.cell(row=total_row_index, column=total_ex_vat_col_index)
    total_ex_vat_total_cell.border = total_ex_vat_border

    # Uncheck gridlines
    ws.sheet_view.showGridLines = False

    # Apply left and bottom borders to all rows except 'Total' and blank row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 2, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row != total_row_index:
                cell.border = border_style

    # Adjust the row height of the blank row above 'Total'
    blank_row_index = total_row_index - 1
    ws.row_dimensions[blank_row_index].height = 7.5

    # Add the 'Updated Data' sheet
    ws_new = wb.create_sheet(title='Updated Data')

    # Write the headers
    for r_idx, col_name in enumerate(df_filtered.columns.tolist(), 1):
        ws_new.cell(row=1, column=r_idx, value=col_name)  # Write headers

    # Write the data
    for r_idx, row in enumerate(df_filtered.values.tolist(), 2):
        for c_idx, value in enumerate(row, 1):
            ws_new.cell(row=r_idx, column=c_idx, value=value)

    # Apply styles to the new sheet
    currency_cols = ['Amount']
    for col in currency_cols:
        if col in df_filtered.columns:
            col_letter = chr(ord('A') + df_filtered.columns.get_loc(col))
            for cell in ws_new[col_letter][1:]:  # Skip header row
                cell.style = currency_format

    for col in numeric_cols:
        if col in df_filtered.columns:
            col_letter = chr(ord('A') + df_filtered.columns.get_loc(col))
            for cell in ws_new[col_letter][1:]:  # Skip header row
                cell.style = number_format

    # Apply blue font formatting for 'DeviceActive' = 'Active'
    device_active_col_index = df_filtered.columns.get_loc('DeviceActive') + 1
    for row in ws_new.iter_rows(min_row=2, max_row=ws_new.max_row, min_col=1, max_col=ws_new.max_column):
        device_active_cell = row[device_active_col_index - 1]  # Indexing correction for 0-based indexing
        if device_active_cell.value == 'Active':
            for cell in row:
                cell.font = blue_font

    # Save the updated workbook with both sheets
    wb.save(excel_path)

    messagebox.showinfo("Success", f"Report successfully saved to {excel_path}")

def on_generate_report():
    """Wrapper function to handle file selection and report generation."""
    csv_file_path = select_file()
    if not csv_file_path:
        messagebox.showwarning("No File Selected", "No CSV file selected. Exiting.")
        return

    excel_path = save_file()
    if not excel_path:
        messagebox.showwarning("No Save Location", "No save location specified. Exiting.")
        return

    process_data(csv_file_path, excel_path)

def main():
    # Create a GUI window
    root = Tk()
    root.title("CSV to Excel Report Generator")
    root.geometry("500x200")  # Set window size
    root.configure(bg="#f0f0f0")  # Set background color

    # Create labels and buttons
    label = Label(root, text="Welcome to the CSV to Excel Report Generator!", font=("Helvetica", 14), bg="#f0f0f0")
    label.pack(pady=20)

    button = Button(root, text="Generate Report", command=on_generate_report, font=("Helvetica", 12), fg="#FFFFFF", bg="#33CC33")
    button.pack(pady=10)

    # Run the GUI event loop
    root.mainloop()

if __name__ == "__main__":
    main()